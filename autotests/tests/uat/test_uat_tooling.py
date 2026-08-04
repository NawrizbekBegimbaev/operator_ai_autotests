from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from autotests.uat.catalog import load_catalog
from autotests.uat.excel import build_workbook, excel_file_name, render_caption_uz
from autotests.uat.i18n_uz import SHEET_TITLES_UZ, missing_translations
from autotests.uat.report import build_report, load_report, render_human_report
from autotests.uat.telegram import TELEGRAM_CAPTION_LIMIT, split_message


JUNIT_TWO_CASES = """<?xml version="1.0" encoding="utf-8"?>
<testsuites><testsuite name="uat" tests="2">
  <testcase name="test[UAT-SYS-001]" time="1.2">
    <properties><property name="uat_id" value="UAT-SYS-001"/></properties>
  </testcase>
  <testcase name="test[UAT-SYS-002]" time="0.3">
    <properties><property name="uat_id" value="UAT-SYS-002"/></properties>
    <failure message="Форма входа не видна">SECRET TRACEBACK</failure>
  </testcase>
</testsuite></testsuites>
"""


def test_uat_report_keeps_three_header_numbers_and_hides_traceback(
    tmp_path: Path,
) -> None:
    junit = tmp_path / "junit.xml"
    junit.write_text(JUNIT_TWO_CASES, encoding="utf-8")
    report = build_report(junit, pytest_returncode=1)
    text = render_human_report(report)
    assert report.passed == 1
    assert report.launched == 2
    assert report.not_ready == 24
    assert report.checkable == 26
    assert report.blocked_defect == 7
    assert report.total == 33
    assert "Пройдено: 1 из 2 запущенных" in text
    assert "24 не готово к запуску" in text
    assert "7 заблокировано дефектом" in text
    assert "всего в чеклисте: 33" in text
    assert "Форма входа не видна" in text
    assert "SECRET TRACEBACK" not in text


def test_telegram_split_preserves_report_and_limits_chunks() -> None:
    report = "\n".join(f"UAT line {index}: " + "x" * 100 for index in range(100))
    chunks = split_message(report, limit=500)
    assert len(chunks) > 1
    assert all(len(chunk) <= 500 for chunk in chunks)
    assert "\n".join(chunks) == report


def test_every_catalog_case_has_uzbek_translation() -> None:
    catalog = load_catalog()
    assert missing_translations([case.id for case in catalog.cases]) == []


def test_excel_report_is_uzbek_and_hides_traceback(tmp_path: Path) -> None:
    junit = tmp_path / "junit.xml"
    junit.write_text(JUNIT_TWO_CASES, encoding="utf-8")
    report = build_report(junit, pytest_returncode=1)

    path = build_workbook(report, tmp_path / "report.xlsx")
    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        SHEET_TITLES_UZ["summary"],
        SHEET_TITLES_UZ["cases"],
        SHEET_TITLES_UZ["defects"],
        SHEET_TITLES_UZ["statuses"],
    ]

    cases = workbook[SHEET_TITLES_UZ["cases"]]
    assert cases.max_row == report.total + 1
    assert cases.cell(row=1, column=5).value == "Nima tekshirildi"
    first_scenario = cases.cell(row=2, column=5).value
    assert first_scenario == "Server ishlayapti va so'rovga javob beradi"
    assert "O'tdi" in str(cases.cell(row=2, column=7).value)
    assert "O'tmadi" in str(cases.cell(row=3, column=7).value)

    dumped = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "SECRET TRACEBACK" not in dumped
    assert "Форма входа не видна" in dumped


def test_telegram_caption_is_uzbek_and_fits_limit(tmp_path: Path) -> None:
    junit = tmp_path / "junit.xml"
    junit.write_text(JUNIT_TWO_CASES, encoding="utf-8")
    report = build_report(junit, pytest_returncode=1)

    caption = render_caption_uz(report)
    assert len(caption) <= TELEGRAM_CAPTION_LIMIT
    assert "Operator AI — kunlik tekshiruv" in caption
    assert "Diqqat talab qiladi" in caption
    assert "SECRET TRACEBACK" not in caption
    assert excel_file_name(report).endswith(".xlsx")


def test_summary_json_round_trip_restores_report(tmp_path: Path) -> None:
    from autotests.uat.report import write_report_files

    junit = tmp_path / "junit.xml"
    junit.write_text(JUNIT_TWO_CASES, encoding="utf-8")
    report = build_report(junit, pytest_returncode=1)
    _, json_path = write_report_files(report, tmp_path / "run")

    assert load_report(json_path) == report
