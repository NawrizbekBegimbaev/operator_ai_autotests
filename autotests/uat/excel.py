"""Excel-отчёт о ежедневном UAT на узбекском языке.

Файл открывает не-IT сотрудник, поэтому:
- первый лист отвечает на один вопрос — «сегодня всё работает или нет»;
- второй лист расшифровывает каждую проверку обычными словами;
- третий лист показывает, что именно ждёт починки разработчиками.
Технические подробности (traceback, логи) в файл не попадают.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from autotests.uat.i18n_uz import (
    COLUMNS_UZ,
    DEFECT_COLUMNS_UZ,
    LEAD_STATUSES_UZ,
    RESULT_HINT_UZ,
    ROLE_COLUMNS_UZ,
    SHEET_TITLES_UZ,
    STATUS_COLUMNS_UZ,
    STATUS_SHEET_NOTES_UZ,
    SUMMARY_LABELS_UZ,
    case_uz,
    duration_uz,
    human_datetime,
    priority_uz,
    role_uz,
    state_with_icon,
    STATE_PLAIN_UZ,
)
from autotests.uat.report import UATReport


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
GREY_FILL = PatternFill("solid", fgColor="E7E6E6")
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

STATE_FILL = {
    "passed": GREEN_FILL,
    "failed": RED_FILL,
    "not_run": YELLOW_FILL,
    "blocked_defect": GREY_FILL,
}


def _write_header(sheet: Worksheet, row: int, columns: tuple[str, ...]) -> None:
    for index, title in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = CELL_BORDER


def _set_widths(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _build_summary_sheet(sheet: Worksheet, report: UATReport) -> None:
    labels = SUMMARY_LABELS_UZ
    failed = sum(item.state == "failed" for item in report.outcomes)
    _set_widths(sheet, (46, 30, 18, 18, 18))

    title = sheet.cell(row=1, column=1, value=labels["title"])
    title.font = TITLE_FONT
    sheet.cell(row=2, column=1, value=labels["generated_at"]).font = LABEL_FONT
    sheet.cell(row=2, column=2, value=human_datetime(report.generated_at))

    verdict_row = 4
    verdict_label = sheet.cell(row=verdict_row, column=1, value=labels["verdict"])
    verdict_label.font = LABEL_FONT
    verdict_text = labels["verdict_green"] if report.is_green else labels["verdict_red"]
    verdict = sheet.cell(row=verdict_row, column=2, value=verdict_text)
    verdict.fill = GREEN_FILL if report.is_green else RED_FILL
    verdict.font = Font(bold=True)
    verdict.alignment = WRAP_TOP

    rows: tuple[tuple[str, object], ...] = (
        (labels["passed"], report.passed),
        (labels["failed"], failed),
        (labels["not_ready"], report.not_ready),
        (labels["blocked"], report.blocked_defect),
        (labels["total"], report.total),
        (labels["duration"], duration_uz(report.duration_seconds)),
    )
    row = verdict_row + 2
    for label, value in rows:
        sheet.cell(row=row, column=1, value=label).font = LABEL_FONT
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value=labels["by_role"]).font = TITLE_FONT
    row += 1
    _write_header(sheet, row, ROLE_COLUMNS_UZ)
    row += 1
    for role in report.roles:
        values = (
            role_uz(role.role),
            role.passed,
            role.launched,
            role.not_ready,
            role.total,
        )
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = CELL_BORDER
            if index == 1:
                cell.alignment = WRAP_TOP
        row += 1

    problems = [item for item in report.outcomes if item.state in {"failed", "not_run"}]
    row += 1
    sheet.cell(row=row, column=1, value=labels["problems"]).font = TITLE_FONT
    row += 1
    if not problems:
        sheet.cell(row=row, column=1, value=labels["no_problems"])
        return
    for item in problems:
        texts = case_uz(item.id)
        cell = sheet.cell(
            row=row,
            column=1,
            value=f"{item.id} — {texts.get('scenario', item.scenario)}",
        )
        cell.alignment = WRAP_TOP
        state_cell = sheet.cell(row=row, column=2, value=state_with_icon(item.state))
        state_cell.fill = STATE_FILL[item.state]
        row += 1


def _build_cases_sheet(sheet: Worksheet, report: UATReport) -> None:
    _set_widths(sheet, (5, 16, 24, 28, 42, 52, 26, 44, 16, 14))
    _write_header(sheet, 1, COLUMNS_UZ)
    # Подсказка при наведении: что означает каждый из четырёх результатов.
    result_header = sheet.cell(row=1, column=7)
    result_header.comment = Comment(RESULT_HINT_UZ, "Operator AI QA", height=140, width=360)
    sheet.freeze_panes = "A2"
    for number, item in enumerate(report.outcomes, start=1):
        row = number + 1
        texts = case_uz(item.id)
        comment = STATE_PLAIN_UZ.get(item.state, "")
        if item.state == "blocked_defect" and item.blocked_by:
            comment = f"{comment} ({', '.join(item.blocked_by)})"
        elif item.state in {"failed", "not_run"} and item.reason:
            comment = f"{comment} {item.reason}"
        values = (
            number,
            item.id,
            role_uz(item.role),
            texts.get("screen", ""),
            texts.get("scenario", item.scenario),
            texts.get("expected", item.expected_result),
            state_with_icon(item.state),
            comment,
            priority_uz(item.priority),
            round(item.duration_seconds, 1),
        )
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = CELL_BORDER
            cell.alignment = WRAP_TOP
            if index == 7:
                cell.fill = STATE_FILL[item.state]
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS_UZ))}{len(report.outcomes) + 1}"


def _build_defects_sheet(sheet: Worksheet, report: UATReport) -> None:
    _set_widths(sheet, (20, 30, 60))
    sheet.cell(row=1, column=1, value=SUMMARY_LABELS_UZ["known_defects"]).font = TITLE_FONT
    _write_header(sheet, 2, DEFECT_COLUMNS_UZ)
    grouped: dict[str, list[str]] = defaultdict(list)
    for item in report.outcomes:
        if item.state != "blocked_defect":
            continue
        for defect in item.blocked_by:
            grouped[defect].append(item.id)
    counts = Counter({defect: len(ids) for defect, ids in grouped.items()})
    row = 3
    for defect, count in sorted(counts.items()):
        values = (defect, count, ", ".join(sorted(grouped[defect])))
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = CELL_BORDER
            cell.alignment = WRAP_TOP
        row += 1


def _build_statuses_sheet(sheet: Worksheet) -> None:
    """Расшифровка статусов лида: в сценариях встречаются Bugun, Chala,
    Ko'tarmadi — без глоссария читателю отчёта они ничего не говорят."""
    _set_widths(sheet, (26, 26, 68, 20, 16, 22))
    sheet.cell(row=1, column=1, value="Lid statuslari — qisqacha izoh").font = TITLE_FONT
    _write_header(sheet, 2, STATUS_COLUMNS_UZ)
    row = 3
    for status in LEAD_STATUSES_UZ:
        values = (
            status["status"],
            status["ru"],
            status["meaning"],
            status["in_queue"],
            status["rank"],
            status["keyword"],
        )
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = CELL_BORDER
            cell.alignment = WRAP_TOP
            if index == 4:
                cell.fill = GREEN_FILL if value == "Ha" else GREY_FILL
        row += 1
    row += 1
    for note in STATUS_SHEET_NOTES_UZ:
        cell = sheet.cell(row=row, column=1, value=f"• {note}")
        cell.alignment = WRAP_TOP
        row += 1


def build_workbook(report: UATReport, path: Path) -> Path:
    """Собрать .xlsx и вернуть путь к нему."""
    workbook = Workbook()
    summary = workbook.active
    assert summary is not None
    summary.title = SHEET_TITLES_UZ["summary"]
    _build_summary_sheet(summary, report)
    _build_cases_sheet(workbook.create_sheet(SHEET_TITLES_UZ["cases"]), report)
    _build_defects_sheet(workbook.create_sheet(SHEET_TITLES_UZ["defects"]), report)
    _build_statuses_sheet(workbook.create_sheet(SHEET_TITLES_UZ["statuses"]))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def excel_file_name(report: UATReport) -> str:
    """Имя файла с датой прогона: в Telegram файлы лежат рядом и должны
    отличаться по названию, а не по порядку загрузки."""
    stamp = report.generated_at[:10]
    return f"Operator-AI-kunlik-hisobot-{stamp}.xlsx"


def render_caption_uz(report: UATReport, *, limit: int = 1024) -> str:
    """Короткая подпись к файлу: Telegram обрезает caption на 1024 символах."""
    labels = SUMMARY_LABELS_UZ
    icon = "🟢" if report.is_green else "🔴"
    failed = sum(item.state == "failed" for item in report.outcomes)
    lines = [
        f"{icon} Operator AI — kunlik tekshiruv",
        f"Sana: {human_datetime(report.generated_at)} (Toshkent)",
        "",
        labels["verdict_green"] if report.is_green else labels["verdict_red"],
        "",
        f"✅ Ishlayapti: {report.passed}",
        f"❌ Muammo bor: {failed}",
        f"⚠️ Tekshirilmadi: {report.not_ready}",
        f"⛔ Ma'lum xatolik sababli to'xtatilgan: {report.blocked_defect}",
        f"Jami ro'yxatda: {report.total}",
        f"Vaqt: {duration_uz(report.duration_seconds)}",
    ]
    problems = [item for item in report.outcomes if item.state in {"failed", "not_run"}]
    if problems:
        lines.extend(["", f"{labels['problems']}:"])
        for item in problems:
            texts = case_uz(item.id)
            lines.append(f"• {item.id} — {texts.get('scenario', item.scenario)}")
    lines.extend(["", "Batafsil ma'lumot biriktirilgan Excel faylda."])
    caption = "\n".join(lines)
    if len(caption) <= limit:
        return caption
    tail = "\n…\nBatafsil ma'lumot biriktirilgan Excel faylda."
    return caption[: limit - len(tail)].rstrip() + tail
